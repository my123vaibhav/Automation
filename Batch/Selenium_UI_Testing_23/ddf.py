import openpyxl

# we need to install 3rd party library----> python present but we use
#file---> workbook--->sheets--->rows--->cells(r,c)

file1="E:\Book1.xlsx"
workbook=openpyxl.load_workbook(file1)
sheet=workbook.active #parent sheet

row=sheet.max_row #5
column=sheet.max_column #3

for r in range(2,row+1):   #range(6)--->   #0(title) to 5
    for c in range(1,column+1):
        if r==2 and c==2:
            sheet.cell(r,c).value='Mumbai'

workbook.save(file1)





import sys
sys.exit(0)
#all row and column se data read
for r in range(2,row+1):   #range(6)--->   #0(title) to 5
    for c in range(1,column+1):
        print(sheet.cell(r,c).value,end='  ')
    print()




import sys

sys.exit(0)
v=sheet.cell(4,2).value #Mumbai
print(v)
v=sheet.cell(1,2).value  #city
print(v)
v=sheet.cell(1,1).value #name
print(v)
v=sheet.cell(2,3).value #89545614
print(v)